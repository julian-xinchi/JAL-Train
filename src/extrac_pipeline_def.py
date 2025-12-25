import argparse
import re
import sys
import inspect
import openpyxl


FULLWIDTH_COMMA = "\uFF0C"
MASTER_HEADER_RE = re.compile(r"^m_.*_\d+$", re.IGNORECASE)
SLAVE_HEADER_RE = re.compile(r"^s_.*_\d+$", re.IGNORECASE)
TRAILING_DIGIT_RE = re.compile(r"_(\d+)$")


def parse_args():
    parser = argparse.ArgumentParser(description="Extract pipeline definitions from Excel.")
    parser.add_argument("excel_file", help="Path to the Excel file")
    parser.add_argument("sheet_name", help="Name of the sheet")
    parser.add_argument("--keyrow", type=int, default=2, help="Row number for master keywords (default: 2)")
    parser.add_argument("--keycol", type=int, default=2, help="Column number for slave keywords (default: 2)")
    parser.add_argument(
        "-e",
        "--errslave",
        action="store_true",
        help="One more row needs to be processed when error slave exists",
    )
    return parser.parse_args()


def sanitize_filename(s):
    return re.sub(r"\W+", "_", s)


def print_boxed(message, *, border_char="@", file=None):
    msg = str(message)
    inner = " " + msg + " "
    top_bottom = border_char * (len(inner) + 2)
    if len(inner) > 64:
        print(msg, file=file)
    else:
        print(top_bottom, file=file)
        print(f"{border_char}{inner}{border_char}", file=file)
        print(top_bottom, file=file)


def fatal(message):
    print_boxed(f"ERROR: {message}", file=sys.stderr)
    raise SystemExit(1)


def split_binary_cell(is_str, *, row, col):
    separator = None
    if "," in is_str:
        separator = ","
    elif "." in is_str:
        separator = "."
    elif FULLWIDTH_COMMA in is_str:
        separator = FULLWIDTH_COMMA

    if separator is not None:
        left, sep, right = is_str.partition(separator)
        if not sep or not left or not right:
            fatal(f"Invalid separator usage in cell at row {row}, col {col}: {is_str} "
                  f"(line {inspect.currentframe().f_back.f_lineno})")
        if not all(c in "01" for c in left) or not all(c in "01" for c in right):   
            fatal(f"Non-binary characters in separated parts at row {row}, col {col}: {is_str} "
                  f"(line {inspect.currentframe().f_back.f_lineno})")
        return left, right

    if not all(c in "01" for c in is_str):
        fatal(f"Invalid format in cell at row {row}, col {col}: {is_str} "
              f"(line {inspect.currentframe().f_back.f_lineno})")

    mid = len(is_str) // 2
    return is_str[:mid], is_str[mid:]


def extract_trailing_digit(header):
    m = TRAILING_DIGIT_RE.search(header)
    if not m:
        return None
    return int(m.group(1))


def find_master_matching(sheet, *, header_row):
    matches = []
    for col in range(1, sheet.max_column + 1):
        hv = sheet.cell(row=header_row, column=col).value
        if hv is not None:
            header = str(hv).strip()
            if MASTER_HEADER_RE.match(header):
                matches.append((col, header))
    return matches


def find_slave_matching(sheet, *, data_start_row, key_col):
    matches = []
    for row in range(data_start_row, sheet.max_row + 1):
        cv = sheet.cell(row=row, column=key_col).value
        if cv is not None:
            header = str(cv).strip()
            if SLAVE_HEADER_RE.match(header):
                matches.append((row, header))
    return matches


def collect_master_data(
    sheet,
    *,
    item_col,
    label,
    data_start_row,
    max_slave_row,
    master_pos,
):
    br_bpp = []
    br_fpp = []
    count = 0

    extract_index = -master_pos
    for rn in range(data_start_row, max_slave_row + 1):
        cur_cell = sheet.cell(row=rn, column=item_col).value
        if cur_cell is None:
            fatal(f"Unexpected empty cell at row {rn}, col {item_col} "
                  f"(line {inspect.currentframe().f_back.f_lineno})")
        cur_cell_nws = re.sub(r"\s+", "", str(cur_cell))
        if cur_cell_nws.upper() in {"NA", "N"}:
            continue
        left, right = split_binary_cell(cur_cell_nws, row=rn, col=item_col)
        if len(left) < master_pos or len(right) < master_pos:
            fatal(
                f"Insufficient bits at row {rn}, col {item_col}: "
                f"need >= {master_pos} per half, got left={len(left)}, right={len(right)} "
                f"(line {inspect.currentframe().f_back.f_lineno})"
            )
        br_bpp.append(right[extract_index])
        br_fpp.append(left[extract_index])
        count += 1

    bpp_tuple = (label, count, "".join(reversed(br_bpp)))
    fpp_tuple = (label, count, "".join(reversed(br_fpp)))
    return bpp_tuple, fpp_tuple


def collect_slave_data(
    sheet,
    *,
    item_row,
    label,
    key_col,
    max_master_col,
    slave_pos1,
    slave_pos2,
):
    br_bpp1 = []
    br_fpp1 = []
    br_bpp2 = []
    br_fpp2 = []
    count = 0

    extract_index1 = -slave_pos1
    extract_index2 = -slave_pos2
    for col in range(key_col + 1, max_master_col + 1):
        cur_cell = sheet.cell(row=item_row, column=col).value
        if cur_cell is None:
            fatal(f"Unexpected empty cell at row {item_row}, col {col} "
                  f"(line {inspect.currentframe().f_back.f_lineno})")
        cur_cell_nws = re.sub(r"\s+", "", str(cur_cell))
        if cur_cell_nws.upper() in {"NA", "N"}:
            continue
        left, right = split_binary_cell(cur_cell_nws, row=item_row, col=col)
        need = max(slave_pos1, slave_pos2)
        if len(left) < need or len(right) < need:
            fatal(
                f"Insufficient bits at row {item_row}, col {col}: "
                f"need >= {need} per half, got left={len(left)}, right={len(right)} "
                f"(line {inspect.currentframe().f_back.f_lineno})"
            )
        br_bpp1.append(right[extract_index1])
        br_fpp1.append(left[extract_index1])
        br_bpp2.append(right[extract_index2])
        br_fpp2.append(left[extract_index2])
        count += 1

    bpp1_tuple = (label, count, "".join(reversed(br_bpp1)))
    fpp1_tuple = (label, count, "".join(reversed(br_fpp1)))
    bpp2_tuple = (label, count, "".join(reversed(br_bpp2)))
    fpp2_tuple = (label, count, "".join(reversed(br_fpp2)))
    return bpp1_tuple, fpp1_tuple, bpp2_tuple, fpp2_tuple


def write_master_block(file, *, bpp_data, fpp_data):
    for label, count, br_fpp_str in fpp_data:
        file.write(f"{f'localparam {label}_BRCH_ARBI_VFPPEN':<40}{f'= {count}{chr(39)}b{br_fpp_str}':<40};\n")
    for label, count, br_bpp_str in bpp_data:
        file.write(f"{f'localparam {label}_BRCH_ARBI_VBPPEN':<40}{f'= {count}{chr(39)}b{br_bpp_str}':<40};\n")
    file.write("\n")
    for label, _, _ in fpp_data:
        file.write(f"{f'localparam {label}_BCH_ARBO_FPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    for label, _, _ in bpp_data:
        file.write(f"{f'localparam {label}_BCH_ARBO_BPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    for label, _, _ in fpp_data:
        file.write(f"{f'localparam {label}_RCH_ARBO_FPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    for label, _, _ in bpp_data:
        file.write(f"{f'localparam {label}_RCH_ARBO_BPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    file.write("\n")


def write_slave_block(
    file,
    *,
    bpp_data1,
    fpp_data1,
    bpp_data2,
    fpp_data2,
):
    for label, count, br_fpp_str in fpp_data1:
        file.write(f"{f'localparam {label}_ACH_ARBI_VFPPEN':<40}{f'= {count}{chr(39)}b{br_fpp_str}':<40};\n")
    for label, count, br_bpp_str in bpp_data1:
        file.write(f"{f'localparam {label}_ACH_ARBI_VBPPEN':<40}{f'= {count}{chr(39)}b{br_bpp_str}':<40};\n")
    file.write("\n")
    for label, _, _ in fpp_data1:
        file.write(f"{f'localparam {label}_ACH_ARBO_FPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    for label, _, _ in bpp_data1:
        file.write(f"{f'localparam {label}_ACH_ARBO_BPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    file.write("\n")
    for label, count, br_fpp_str in fpp_data2:
        file.write(f"{f'localparam {label}_WCH_ARBI_VFPPEN':<40}{f'= {count}{chr(39)}b{br_fpp_str}':<40};\n")
    for label, count, br_bpp_str in bpp_data2:
        file.write(f"{f'localparam {label}_WCH_ARBI_VBPPEN':<40}{f'= {count}{chr(39)}b{br_bpp_str}':<40};\n")
    file.write("\n")
    for label, _, _ in fpp_data2:
        file.write(f"{f'localparam {label}_WCH_ARBO_FPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    for label, _, _ in bpp_data2:
        file.write(f"{f'localparam {label}_WCH_ARBO_BPPEN':<40}{f'= 1{chr(39)}b0':<40};\n")
    file.write("\n")


def main():
    args = parse_args()

    master_pos = 1
    slave_pos1 = 1
    slave_pos2 = 1

    try:
        workbook = openpyxl.load_workbook(args.excel_file)
    except Exception as exc:
        fatal(f"Failed to open Excel file '{args.excel_file}': {exc}")
    try:
        if args.sheet_name not in workbook.sheetnames:
            fatal(f"Sheet '{args.sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames} "
                  f"(line {inspect.currentframe().f_back.f_lineno})")
        sheet = workbook[args.sheet_name]

        header_row = args.keyrow
        key_col = args.keycol
        data_start_row = header_row + 1

        master_matching = find_master_matching(sheet, header_row=header_row)
        slave_matching = find_slave_matching(sheet, data_start_row=data_start_row, key_col=key_col)

        max_master_col = max((col for col, _ in master_matching), default=0)
        max_slave_row = max((row for row, _ in slave_matching), default=0)
        if args.errslave:
            max_slave_row = min(max_slave_row + 1, sheet.max_row)

        out_fname = f"{sanitize_filename(args.sheet_name)}_def.txt"

        if not master_matching:
            print("Warning: No matching items found for master mode.")
        if not slave_matching:
            print("Warning: No matching items found for slave mode.")

        if master_matching or slave_matching:
            with open(out_fname, "w") as file:
                if master_matching:
                    bpp_data = []
                    fpp_data = []
                    for idx, (item_col, item_header) in enumerate(master_matching):
                        digit = extract_trailing_digit(item_header)
                        if digit is None:
                            fatal(f"Invalid master header (missing trailing digit): {item_header} "
                                  f"(line {inspect.currentframe().f_back.f_lineno})")
                        if digit != idx:
                            fatal(f"Master header index mismatch: expected _{idx}, got _{digit} in {item_header} "
                                  f"(line {inspect.currentframe().f_back.f_lineno})")
                        label = f"M{idx}"
                        bpp_tuple, fpp_tuple = collect_master_data(
                            sheet,
                            item_col=item_col,
                            label=label,
                            data_start_row=data_start_row,
                            max_slave_row=max_slave_row,
                            master_pos=master_pos,
                        )
                        bpp_data.append(bpp_tuple)
                        fpp_data.append(fpp_tuple)
                    write_master_block(file, bpp_data=bpp_data, fpp_data=fpp_data)

                if slave_matching:
                    bpp_data1 = []
                    fpp_data1 = []
                    bpp_data2 = []
                    fpp_data2 = []
                    for idx, (item_row, item_header) in enumerate(slave_matching):
                        digit = extract_trailing_digit(item_header)
                        if digit is None:
                            fatal(f"Invalid slave header (missing trailing digit): {item_header} "
                                  f"(line {inspect.currentframe().f_back.f_lineno})")
                        if digit != idx:
                            fatal(f"Slave header index mismatch: expected _{idx}, got _{digit} in {item_header} "
                                  f"(line {inspect.currentframe().f_back.f_lineno})")
                        label = f"S{idx}"
                        bpp1, fpp1, bpp2, fpp2 = collect_slave_data(
                            sheet,
                            item_row=item_row,
                            label=label,
                            key_col=key_col,
                            max_master_col=max_master_col,
                            slave_pos1=slave_pos1,
                            slave_pos2=slave_pos2,
                        )
                        bpp_data1.append(bpp1)
                        fpp_data1.append(fpp1)
                        bpp_data2.append(bpp2)
                        fpp_data2.append(fpp2)
                    write_slave_block(
                        file,
                        bpp_data1=bpp_data1,
                        fpp_data1=fpp_data1,
                        bpp_data2=bpp_data2,
                        fpp_data2=fpp_data2,
                    )

        print_boxed(f"Contents written to file: {out_fname}")
        return 0
    except SystemExit:
        raise
    except Exception as exc:
        fatal(f"Unhandled error: {type(exc).__name__}: {exc}")
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
