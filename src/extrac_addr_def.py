
import sys
import openpyxl
import re

# Check that enough command-line arguments are provided
if len(sys.argv) < 3:
    print("Usage: python script.py excel_file sheet_name start_cell end_cell")
    sys.exit(1)

# Get Excel filename, sheet name, start cell and end cell addresses from command-line arguments
excel_file = sys.argv[1]
sheet_name = sys.argv[2]

# Open the Excel file and the specified worksheet
workbook = openpyxl.load_workbook(excel_file)
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    print(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
    sys.exit(1)

# Initialize an empty list to store row numbers
master_rows = []

# Iterate column A to find rows with value "master" (case-insensitive)
for row in sheet['A']:
    if row.value and row.value.lower() == "master":
        master_rows.append(row.row)

# Initialize an empty list to store row numbers
slave_rows = []

# Iterate column A to find rows with value "slave" (case-insensitive)
for row in sheet['A']:
    if row.value and row.value.lower() == "slave":
        slave_rows.append(row.row)

header_row = 2
data_start_row = 3
header_re = re.compile(r'^s_.*_\d+$', re.IGNORECASE)          # match "s_*_\\d"
hex_re = re.compile(r'^0[xX][0-9A-Fa-f]{4}_[0-9A-Fa-f]{4}$')     # match "0xxxxx_xxxx" hex format
hex1_re = re.compile(r'^[0-9A-Fa-f]{4}_[0-9A-Fa-f]{4}$')     # match "xxxx_xxxx" hex format
hex2_re = re.compile(r'^(0[xX])?([0-9A-Fa-f]{8})$')     # match "0xHHHHHHHH" or "HHHHHHHH" hex format

# find matching columns by header in row 2
matching_cols = []
for col in range(1, sheet.max_column + 1):
    hv = sheet.cell(row=header_row, column=col).value
    if hv and header_re.match(str(hv).strip()):
        matching_cols.append((col, str(hv).strip()))

# output filename based on sheet_name (sanitize non-word chars)
out_fname = re.sub(r'\W+', '_', sheet_name) + '_def.v'

# append processed data to params.txt
if matching_cols:
    with open(out_fname, 'w') as file:
        for col_idx, header in matching_cols:
            # extract trailing digit(s) for S\d
            m = re.search(r'_(\d+)$', header)
            s_digit = m.group(1) if m else ''

            # collect values from row 3 onwards, stop on "NA" (not recorded)
            vals = []
            r = data_start_row
            while r <= sheet.max_row:
                cv = sheet.cell(row=r, column=col_idx).value
                if cv is None:
                    break
                s = str(cv).strip()
                if s.upper() == 'NA':
                    break
                if hex_re.match(s) or hex1_re.match(s):
                    vals.append(s)
                elif hex2_re.match(s):
                    prefix, hex_part = hex2_re.match(s).groups()
                    formatted = hex_part[:4] + '_' + hex_part[4:]  # format as xxxx_xxxx
                    vals.append(formatted)
                r += 1

            # every two values form a group; ignore dangling last if odd count
            for grp_i in range(0, len(vals) - 1, 2):
                data0 = vals[grp_i]
                data1 = vals[grp_i + 1]
                group_no = grp_i // 2
                # remove leading 0x/0X if present
                data0_clean = re.sub(r'^0[xX]', '', data0)
                data1_clean = re.sub(r'^0[xX]', '', data1)
                file.write(f"{f'localparam S{s_digit}_REGION{group_no}_SA ':<40}{f'= 32\'h{data0_clean}':<40};\n")
                file.write(f"{f'localparam S{s_digit}_REGION{group_no}_EA ':<40}{f'= 32\'h{data1_clean}':<40};\n")

else:
    print(f"No matching columns found in sheet '{sheet_name}'")

# Close the workbook
workbook.close()

print("Contents written to file:", out_fname)
print("Done")
