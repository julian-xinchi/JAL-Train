
import sys
import openpyxl
import re

# Check that enough command-line arguments are provided
if len(sys.argv) < 3:
    print("Usage: python script.py excel_file sheet_name start_cell end_cell")
    sys.exit(1)

# Get Excel filename, sheet name, start cell and end cell addresses from command-line arguments
excel_file = sys.argv[1]
sheet_name = sys.argv[2]
# start_cell = sys.argv[3]
# end_cell = sys.argv[4]

# Open the Excel file and the specified worksheet
workbook = openpyxl.load_workbook(excel_file)
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    print(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
    sys.exit(1)

# Calculate start and end row/column numbers
# start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)  # e.g. A1 -> ('A', 1)
# end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)  # e.g. A5 -> ('A', 5)

# Initialize an empty list to store row numbers
master_rows = []

# Iterate column A to find rows with value "master" (case-insensitive)
for row in sheet['A']:
    if row.value and row.value.lower() == "master":
        master_rows.append(row.row)

# Initialize an empty list to store row numbers
slave_rows = []

# Iterate column A to find rows with value "slave" (case-insensitive)
for row in sheet['A']:
    if row.value and row.value.lower() == "slave":
        slave_rows.append(row.row)

# output filename based on sheet_name (sanitize non-word chars)
out_fname = re.sub(r'\W+', '_', sheet_name) + '_otn_def.v'

# Open a text file and write contents
with open(out_fname, 'w') as file:
    i = 0
    for row in master_rows:
        cell_address = f"O{row}"
        cell_value = sheet[cell_address].value
        file.write(f"{f'localparam M{i}_OTN ':<40}{f'= {cell_value}':<40};\n")
        i = i + 1
    i = 0
    for row in master_rows:
        cell_address = f"O{row}"
        cell_value = sheet[cell_address].value
        try:
            val = int(cell_value) // 2
        except (TypeError, ValueError):
            val = cell_value
        file.write(f"{f'localparam M{i}_WOTN ':<40}{f'= {val}':<40};\n")
        i = i + 1
    i = 0
    for row in slave_rows:
        cell_address = f"O{row}"
        cell_value = sheet[cell_address].value
        file.write(f"{f'localparam S{i}_OTN ':<40}{f'= {cell_value}':<40};\n")
        i = i + 1
    i = 0
    for row in slave_rows:
        cell_address = f"O{row}"
        cell_value = sheet[cell_address].value
        try:
            val = int(cell_value) // 2
        except (TypeError, ValueError):
            val = cell_value
        file.write(f"{f'localparam S{i}_WOTN ':<40}{f'= {val}':<40};\n")
        i = i + 1

# Close the workbook
workbook.close()

print("Contents written to file:", out_fname)
print("Done")
